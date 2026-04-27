from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import os

def create_api_parameters_excel():
    """创建API接口数据传输参数Excel表格"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "API接口数据传输参数"
    
    # 设置标题样式
    title_font = Font(name='微软雅黑', size=14, bold=True, color='FFFFFF')
    header_font = Font(name='微软雅黑', size=11, bold=True)
    normal_font = Font(name='微软雅黑', size=10)
    
    title_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    
    # 添加主标题
    ws.merge_cells('A1:H1')
    ws['A1'] = '传感器数据服务器 - API接口数据传输参数表'
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # 添加表头
    headers = [
        '接口名称',
        '请求方法',
        'URL路径',
        '参数名称',
        '参数类型',
        '是否必填',
        '传输频率',
        '说明'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # 设置列宽
    column_widths = [20, 12, 35, 20, 12, 10, 15, 40]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # API接口数据
    api_data = [
        # 健康检查接口
        ['健康检查', 'GET', '/api/health', '-', '-', '-', '-', '返回服务状态和版本信息'],
        
        # 统计信息接口
        ['统计信息', 'GET', '/api/stats', '-', '-', '-', '-', '返回系统运行统计数据'],
        
        # 通用数据接收接口
        ['数据接收', 'POST', '/api/receive', 'sensor_type', 'string', '是', '实时', '传感器类型（skin/environment/device）'],
        ['', '', '', 'data', 'object', '是', '实时', '传感器数据对象'],
        ['', '', '', 'timestamp', 'string', '否', '实时', '数据采集时间戳'],
        ['', '', '', 'device_id', 'string', '否', '实时', '设备唯一标识'],
        
        # 皮肤传感器专用接口
        ['皮肤传感器', 'POST', '/api/sensor/skin', 'moisture', 'int', '是', '2秒/次', '皮肤湿度（0-100）'],
        ['', '', '', 'oiliness', 'int', '是', '2秒/次', '皮肤油脂度（0-100）'],
        ['', '', '', 'temperature', 'float', '否', '2秒/次', '皮肤温度（℃）'],
        ['', '', '', 'device_id', 'string', '是', '2秒/次', '设备ID'],
        ['', '', '', 'timestamp', 'string', '是', '2秒/次', '采集时间'],
        
        # 环境传感器专用接口
        ['环境传感器', 'POST', '/api/sensor/environment', 'temperature', 'float', '是', '3秒/次', '环境温度（℃）'],
        ['', '', '', 'humidity', 'float', '是', '3秒/次', '环境湿度（%）'],
        ['', '', '', 'pm25', 'int', '是', '3秒/次', 'PM2.5浓度（μg/m³）'],
        ['', '', '', 'co2', 'int', '是', '3秒/次', 'CO₂浓度（ppm）'],
        ['', '', '', 'light_intensity', 'float', '否', '3秒/次', '光照强度（lux）'],
        ['', '', '', 'noise_level', 'float', '否', '3秒/次', '噪音等级（dB）'],
        ['', '', '', 'location', 'object', '否', '3秒/次', '地理位置{lat, lng}'],
        ['', '', '', 'device_id', 'string', '是', '3秒/次', '设备ID'],
        ['', '', '', 'timestamp', 'string', '是', '3秒/次', '采集时间'],
        
        # 设备状态接口
        ['设备状态', 'POST', '/api/device/status', 'device_id', 'string', '是', '30秒/次', '设备唯一标识'],
        ['', '', '', 'status', 'string', '是', '30秒/次', '设备状态（online/offline/error）'],
        ['', '', '', 'battery_level', 'int', '是', '30秒/次', '电池电量（0-100）'],
        ['', '', '', 'signal_strength', 'int', '是', '30秒/次', '信号强度（-100~0 dBm）'],
        ['', '', '', 'firmware_version', 'string', '是', '30秒/次', '固件版本'],
        ['', '', '', 'last_heartbeat', 'string', '是', '30秒/次', '最后心跳时间'],
        ['', '', '', 'error_code', 'string', '否', '30秒/次', '错误代码'],
        ['', '', '', 'uptime', 'int', '否', '30秒/次', '运行时长（秒）'],
        ['', '', '', 'memory_usage', 'float', '否', '30秒/次', '内存使用率（%）'],
        
        # JWT认证接口
        ['JWT登录', 'POST', '/api/auth/login', 'username', 'string', '是', '按需', '用户名'],
        ['', '', '', 'password', 'string', '是', '按需', '密码'],
        
        # API Key生成接口
        ['生成API Key', 'POST', '/api/auth/apikey', 'device_name', 'string', '是', '按需', '设备名称'],
        ['', '', '', 'description', 'string', '否', '按需', '密钥描述'],
        
        # 加密数据接收接口
        ['加密数据接收', 'POST', '/api/receive/secure', 'encrypted_data', 'string', '是', '实时', 'AES加密后的数据'],
        ['', '', '', 'api_key', 'string', '是', '实时', 'API密钥'],
        
        # API Key认证接收接口
        ['API Key认证接收', 'POST', '/api/receive/apikey', 'sensor_type', 'string', '是', '实时', '传感器类型'],
        ['', '', '', 'data', 'object', '是', '实时', '传感器数据'],
        ['', '', '', 'X-API-Key', 'string', '是', '实时', 'HTTP Header中的API Key'],
        
        # 数据加密接口
        ['数据加密', 'POST', '/api/encrypt', 'data', 'object', '是', '按需', '待加密的JSON数据'],
        ['', '', '', 'api_key', 'string', '是', '按需', 'API密钥'],
        
        # 数据解密接口
        ['数据解密', 'POST', '/api/decrypt', 'encrypted_data', 'string', '是', '按需', 'Base64编码的加密数据'],
        ['', '', '', 'api_key', 'string', '是', '按需', 'API密钥'],
    ]
    
    # 填充数据
    row_num = 3
    for data in api_data:
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.font = normal_font
            cell.alignment = Alignment(vertical='center', wrap_text=True)
        row_num += 1
    
    # 添加边框
    from openpyxl.styles import Border, Side
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=8):
        for cell in row:
            cell.border = thin_border
    
    # 冻结首行
    ws.freeze_panes = 'A3'
    
    # 保存文件
    output_path = os.path.join(os.path.dirname(__file__), 'docs', 'API接口数据传输参数表.xlsx')
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    
    print(f"✅ API接口数据传输参数表已生成：{output_path}")
    return output_path

if __name__ == '__main__':
    create_api_parameters_excel()
